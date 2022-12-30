import os
import pandas as pd
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from datetime import datetime
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup


class Stock_report():
    def __init__(self):
        global url_goo, url_yah, f_name, save_path, browser
        url_goo = 'https://www.google.com/search?q='
        url_yah = 'https://finance.yahoo.com/quote/'
        f_name = '../report/stock_report.xlsx'
        save_path = f"./test.xlsx"
        browser = webdriver.Chrome()
        self.stock = []
        self.stock_etf = []
        self.prices_now = []
        self.prices_52w = []
        self.prices_mon = []
        self.per_today = []
        self.d_alarms = f''
        self.n_alarms = f''

    def process(self):
        self.stock_req()
        self.mon_checker()
        self.crawl_prices()
        self.update_file()
        self.alarm()

    def stock_req(self):
        # stock = input('Please tell me tickers of stocks to crawl. : ').split()
        self.stock = ['aapl',
                       'googl', 'nvda',
                       'tsla', 'ko',
                       'pep',
                       'asml']
        self.stock_etf = ['qqq',
                           'spy']

    def mon_checker(self):
        wb = openpyxl.load_workbook(f_name)
        ws = wb.active
        if str(ws['B2'].value) != (str(datetime.today())[5:7]+'월'):
            ws['B2'].value = (str(datetime.today())[5:7]+'월')
            ws['B14'].value = (str(datetime.today())[5:7] + '월')
            for i in (self.stock + self.stock_etf):
                browser.get(url_yah + i + '/history?p=' + i)
                date_choice = str(datetime.today().strftime("%b")) + ' 01, ' + str(datetime.today().strftime("%Y"))
                soup = BeautifulSoup(browser.page_source, features="lxml")
                price_mon = soup.find('span', text=date_choice).parent.next_sibling.text
                self.prices_mon.append(float(price_mon))
                browser.quit()
            for i,j in enumerate(self.prices_mon):
                ws["C"+str(15 + i)].value = j
                wb.save(f_name)
        print('checking month is completed.')

    def crawl_prices(self):
        for i in (self.stock + self.stock_etf):
            browser.get(url_goo + i + '+stock')
            soup = BeautifulSoup(browser.page_source, features="lxml")
            find_now = soup.find_all('span', attrs={'class': 'IsqQVc NprOob wT3VGc'})
            text_nows = [float(i.text) for i in find_now]
            find_52w = soup.find_all('div', attrs={'data-attrid': '52-주 최고'})
            text_52ws = [float(i.text) for i in find_52w]
            find_per = soup.find('span', attrs={'class': 'jBBUv'})
            text_per = float(find_per.text[1:-2])
            self.prices_now += text_nows
            self.prices_52w += text_52ws
            self.per_today.append(text_per)
        browser.quit()
        print('crawling is completed.')

    def update_file(self):
        wb = openpyxl.load_workbook(f_name)
        ws = wb.active
        for i,j in enumerate(self.prices_now):
            ws["E"+str(15 + i)].value = j
        for i,j in enumerate(self.prices_52w):
            ws["C"+str(3 + i)].value = j
        wb.save(save_path)
        print('updating file is completed.')


    def alarm(self):
        wb = openpyxl.load_workbook(save_path, data_only=True)
        ws = wb.active

        d_alarm = [f'{self.stock[i]}: {j}%    ' if abs(j) > 5 else '' for i,j in enumerate(self.per_today)]
        self.d_alarms = ''.join(d_alarm)
        print(self.d_alarms)

        for i, j in enumerate(self.per_today):
            isnode = ws['F' + str(15 + i)].value * 100
            print(isnode)
            # if str(isnode - j)[1] != str(isnode)[1]:
            #     n_alarm = f'{self.stock[i]}: {int(isnode)}%    '
            #     self.n_alarms += n_alarm

        if (self.d_alarms != '') or (self.n_alarms != ''):
            wb = openpyxl.load_workbook(save_path)
            ws = wb.active
            ws["M15"].value = self.d_alarms


if __name__ == '__main__':
    sr = Stock_report()
    sr.process()
