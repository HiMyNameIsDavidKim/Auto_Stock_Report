import os
import pandas as pd
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from datetime import datetime
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup


class Stock_report():
    def __init__(self):
        self.url_goo = 'https://www.google.com/search?q='
        self.url_yah = 'https://finance.yahoo.com/quote/'
        self.stock = []
        self.stock_etf = []
        self.prices_now = []
        self.prices_52w = []
        self.prices_mon = []
        self.f_name = './report/stock_report.xlsx'
        self.browser = webdriver.Chrome()

    def process(self):
        self.stock_req()
        self.crawl_prices()
        self.update_file()

    def stock_req(self):
        # stock = input('Please tell me the stock to crawl. : ').split()
        self.stock = ['aapl',
                       'googl', 'nvda',
                       'tsla', 'ko',
                       'pep',
                       'asml']
        self.stock_etf = ['qqq',
                           'spy']

    def crawl_prices(self):
        browser = self.browser
        stock = self.stock
        stock_etf = self.stock_etf
        prices_mon = self.prices_mon

        for i in (stock + stock_etf):
            browser.get(self.url_goo + i + '+stock')
            soup = BeautifulSoup(browser.page_source, features="lxml")
            find_now = soup.find_all('span', attrs={'class': 'IsqQVc NprOob wT3VGc'})
            text_nows = [float(i.text) for i in find_now]
            find_52w = soup.find_all('div', attrs={'data-attrid': '52-주 최고'})
            text_52ws = [float(i.text) for i in find_52w]
            self.prices_now += text_nows
            self.prices_52w += text_52ws

        for i in (stock + stock_etf):
            browser.get(self.url_yah + i + '/history?p=' + i)
            date_choice = str(datetime.today().strftime("%b")) + ' 01, ' + str(datetime.today().strftime("%Y"))
            soup = BeautifulSoup(browser.page_source, features="lxml")
            price_mon = soup.find('span', text=date_choice).parent.next_sibling.text
            prices_mon.append(float(price_mon))

        browser.quit()

    def update_file(self):
        f_name = self.f_name
        prices_now = self.prices_now
        prices_52w = self.prices_52w
        prices_mon = self.prices_mon
        wb = openpyxl.load_workbook(f_name)
        ws = wb.active
        for i,j in enumerate(prices_now):
            ws["E"+str(15 + i)].value = j
        for i,j in enumerate(prices_52w):
            ws["C"+str(3 + i)].value = j
        if str(ws['B2'].value + '월') != (str(datetime.today())[5:7]+'월'):
            ws['B2'].value = (str(datetime.today())[5:7]+'월')
            ws['B14'].value = (str(datetime.today())[5:7] + '월')
            for i,j in enumerate(prices_mon):
                ws["C"+str(15 + i)].value = j
        wb.save(f"./report/stock_report_{str(datetime.today())[2:11]}.xlsx")


if __name__ == '__main__':
    sr = Stock_report()
    sr.process()
