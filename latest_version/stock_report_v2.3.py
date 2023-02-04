import os
import pandas as pd
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from datetime import datetime
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import yfinance as yf


class Stock_report():
    def __init__(self):
        global url_goo, url_yah, url_fed, url_cpi, url_yoy, \
            f_name, save_path, browser
        url_goo = 'https://www.google.com/search?q='
        url_fed = 'https://kr.investing.com/central-banks/fed-rate-monitor'
        url_cpi = 'https://www.bls.gov/schedule/news_release/cpi.htm'
        url_yoy = 'https://www.investing.com/economic-calendar/cpi-733'
        f_name = '../report/stock_report.xlsx'
        save_path = f"../report/stock_report_{str(datetime.today())[2:11]}.xlsx"
        browser = webdriver.Chrome()
        self.stock = []
        self.stock_etf = []
        self.prices_now = []
        self.prices_52w = []
        self.prices_mon = []
        self.per_today = []
        self.fed_date = None
        self.cpi_date = None
        self.d_alarms = f''
        self.n_alarms = f''
        self.f_alarms = f''
        self.c_alarms = f''

    def process(self):
        self.stock_req()
        self.mon_checker()
        self.crawl_prices()
        self.update_file()
        self.alarm()

    def stock_req(self):
        # stock = input('Please tell me tickers of stocks to crawl. : ').split()
        self.stock = ['AAPL',
                       'KO', 'NVDA',
                       'TSLA', 'GOOGL',
                       'PEP',
                       'ASML']
        self.stock_etf = ['qqq',
                           'spy']

    def mon_checker(self):
        wb = openpyxl.load_workbook(f_name)
        ws = wb.active
        if str(ws['B2'].value) != (str(datetime.today())[5:7]+'M'):
            ws['B2'].value = (str(datetime.today())[5:7]+'M')
            ws['B14'].value = (str(datetime.today())[5:7] + 'M')
            for i in (self.stock + self.stock_etf):
                df = yf.download(i, start=str(datetime.today())[:9] + '1', end=str(datetime.today())[:9] + '2')
                price_mon = df['Close'].sum()
                self.prices_mon.append(float(price_mon))
            for i, j in enumerate(self.prices_mon):
                ws["C"+str(15 + i)].value = j
                wb.save(f_name)
        print('### Checking month is completed. ###')

    def crawl_prices(self):
        for i in (self.stock + self.stock_etf):
            browser.get(url_goo + i + '+stock')
            soup = BeautifulSoup(browser.page_source, features="lxml")
            find_now = soup.find_all('span', attrs={'class': 'IsqQVc NprOob wT3VGc'})
            text_nows = [float(i.text) for i in find_now]
            find_52w = soup.find_all('div', attrs={'data-attrid': '52-주 최고'})
            text_52ws = [float(i.text) for i in find_52w]
            self.prices_now += text_nows
            self.prices_52w += text_52ws
            find_per = soup.find('span', attrs={'class': 'jBBUv'})
            text_per = float(find_per.text[1:-2])
            self.per_today.append(text_per)
        print('### Crawling price is completed. ###')

        browser.get(url_fed)
        soup = BeautifulSoup(browser.page_source, features="lxml")
        self.fed_date = soup.find('div', attrs={'class': 'fedRateDate'}).text
        print('### Crawling FED interest date is completed. ###')

        browser.get(url_cpi)
        df = pd.read_html(browser.page_source)[1]
        self.cpi_date = [i for i in df['Release Date']]
        print('### Crawling cpi date is completed. ###')

        browser.quit()


    def update_file(self):
        wb = openpyxl.load_workbook(f_name)
        ws = wb.active
        for i,j in enumerate(self.prices_now):
            ws["E"+str(15 + i)].value = j
        for i,j in enumerate(self.prices_52w):
            ws["C"+str(3 + i)].value = j
        wb.save(save_path)
        print('### Updating file is completed. ###')

    def alarm(self):
        wb = openpyxl.load_workbook(save_path)
        ws = wb.active

        d_alarm = [f'{self.stock[i]} has a big change. {j}%    '
                   if abs(j) > 5 else ''
                   for i, j in enumerate(self.per_today)]
        self.d_alarms = ''.join(d_alarm)

        # for i, j in enumerate(self.per_today[:7]):
        #     rate_today = (ws['E' + str(15 + i)].value - ws['C' + str(3 + i)].value) / ws['C' + str(3 + i)].value
        #     isnode = round(rate_today * 100)
        #     if str(isnode - j)[1] != str(isnode)[1]:
        #         n_alarm = f'{self.stock[i]} is arrived new node. {int(isnode)}%    '
        #         self.n_alarms += n_alarm

        target_date = datetime(int(self.fed_date[:5]), int(self.fed_date[7:9]), int(self.fed_date[11:13]))
        d_day = target_date - datetime.today()
        if d_day.days < 8:
            self.f_alarms = f"FED's interest rate announcement is {d_day.days} days away." # url_fed

        for cpi_date in self.cpi_date:
            target_date = datetime(int(cpi_date[-4:]), int(self.month_to_num(cpi_date[:3])), int(cpi_date[-8:-6]))
            d_day = target_date - datetime.today()
            if 0 < d_day.days < 8:
                self.c_alarms = f"CPI announcement is {d_day.days} days away." # url_yoy

        if (self.d_alarms != '') or (self.n_alarms != ''):
            wb = openpyxl.load_workbook(save_path)
            ws = wb.active
            ws["M15"].value = self.d_alarms
            ws["M16"].value = self.n_alarms
            ws["M17"].value = self.f_alarms
            ws["M18"].value = self.c_alarms
        wb.save(save_path)
        print('### Checking alarm is completed. ###')

    def month_to_num(self, month):
        return {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}[month]


if __name__ == '__main__':
    sr = Stock_report()
    sr.process()