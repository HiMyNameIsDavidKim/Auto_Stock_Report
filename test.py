import os
import pandas as pd
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from datetime import datetime


class Stock_report():
    def __init__(self):
        self.url = 'https://www.investing.com/equities/'
        self.url_etf = 'https://www.investing.com/etfs/'
        self.stock = []
        self.stock_etf = []
        self.prices_now = []
        self.prices_52w = []
        self.prices_mon = []
        self.f_name = './report/stock_report.xlsx'
        self.browser = webdriver.Chrome()

    def process(self):
        self.stock_req()
        self.crawl_prices()
        self.update_file()

    def stock_req(self):
        # stock = input('Please tell me the ticker of the stock to crawl. : ').split()
        self.stock = ['apple-computer-inc',
                       'google-inc', 'nvidia-corp',
                       'tesla-motors', 'coca-cola-co',
                       'pepsico',
                       'asml-holdings']
        self.stock_etf = ['powershares-qqqq',
                           'proshares-ultra-qqq-etf',
                           'spdr-s-p-500']

    def crawl_prices(self):
        browser = self.browser
        stock = self.stock
        stock_etf = self.stock_etf
        prices_now = self.prices_now
        prices_52w = self.prices_52w
        prices_mon = self.prices_mon
        xp_now = '//*[@id="__next"]/div/div/div/div[2]/main/div/div[1]/div[2]/div[1]/span'
        xp_52w = '//*[@id="__next"]/div/div/div/div[2]/main/div/div[1]/div[2]/ul/li[3]/div[2]'
        for i in stock:
            browser.get(self.url + i)
            price_now = browser.find_element(By.XPATH, xp_now).text
            prices_now.append(float(price_now))
            price_52w = browser.find_element(By.XPATH, xp_52w).text
            price_52w = price_52w[(price_52w.find('-') + 2):]
            prices_52w.append(float(price_52w))
        xp_now = '//*[@id="last_last"]'
        xp_52w = '//*[@id="leftColumn"]/div[9]/div[5]/span[2]'
        for i in stock_etf:
            browser.get(self.url_etf + i)
            price_now = browser.find_element(By.XPATH, xp_now).text
            prices_now.append(float(price_now))
            price_52w = browser.find_element(By.XPATH, xp_52w).text
            price_52w = price_52w[(price_52w.find('-') + 2):]
            prices_52w.append(float(price_52w))
        for i in stock:
            browser.get(self.url + i + '-historical-data')
            date_choice = str(datetime.today())[5:7] + '/01/' + str(datetime.today())[:4]
            df = pd.read_html(browser.page_source)[1]
            price_mon = df.loc[df['Date'] == date_choice]['Price']
            prices_mon.append(float(price_mon))
        browser.quit()

    def update_file(self):
        f_name = self.f_name
        prices_now = self.prices_now
        prices_52w = self.prices_52w
        prices_mon = self.prices_mon
        wb = openpyxl.load_workbook(f_name)
        ws = wb.active
        for i,j in enumerate(prices_now):
            ws["E"+str(15 + i)].value = j
        for i,j in enumerate(prices_52w):
            ws["C"+str(3 + i)].value = j
        if str(ws['B2'].value + '월') != (str(datetime.today())[5:7]+'월'):
            ws['B2'].value = (str(datetime.today())[5:7]+'월')
            ws['B14'].value = (str(datetime.today())[5:7] + '월')
            for i,j in enumerate(prices_mon):
                ws["C"+str(15 + i)].value = j
        wb.save(f"./report/stock_report_{str(datetime.today())[2:11]}.xlsx")


if __name__ == '__main__':
    sr = Stock_report()
    sr.process()
